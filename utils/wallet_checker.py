import asyncio
from termcolor import cprint, colored
import os
import time
import random
import pandas as pd
import decimal

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from utils.networks import HyperTestnetRPC
from utils.multicall3 import Multicall3
from modules import Client
from dev import GeneralSettings
from config import (
    TOKENS_PER_CHAIN, ACCOUNTS_DATA,
    MULTICALL3_CONTRACTS, MULTICALL3_ABI
)


def get_accounts():
    accounts_dict = ACCOUNTS_DATA['accounts']
    
    account_names = []
    private_keys = []
    proxies = []
    
    for account_name, account_data in accounts_dict.items():
        account_names.append(account_name)
        private_keys.append(account_data.get('evm_private_key'))
        # Assuming there might be a proxy field, add None if not present
        proxies.append(account_data.get('proxy', None))
    
    return account_names, private_keys, proxies

ACCOUNT_NAMES, PRIVATE_KEYS, PROXIES = get_accounts()

class MulticallTxChecker:
    def __init__(self, batch_size=50):
        """Init with batch size."""
        self.batch_size = batch_size

    async def check_wallets(self):
        """Main entry: run multicall in batches. If fail, halve batch size and retry."""
        start_time = time.time()
        cprint('✅ Processing wallets...', 'green')
        accounts_data = self.get_accounts_data()

        current_batch_size = self.batch_size
        while current_batch_size > 0:
            try:
                batched_accounts = list(self.chunk_list(accounts_data, current_batch_size))
                
                batch_tasks = [
                    self.process_batch(batch, idx) 
                    for idx, batch in enumerate(batched_accounts, start=1)
                ]
                batch_results = await asyncio.gather(*batch_tasks)
                
                all_wallets_data = []
                for results in batch_results:
                    all_wallets_data.extend(results)

                await self.save_to_excel(all_wallets_data)
                execution_time = time.time() - start_time
                cprint(f'⏱️ Total execution time for {len(all_wallets_data)} wallets: '
                       f'{execution_time:.2f} seconds', 'light_blue')
                return

            except Exception as e:
                # If there's an error, halve the batch size and retry
                cprint(f"❌ Error in multicall: {e}", 'red')
                new_batch_size = current_batch_size // 2
                cprint(f"Retrying with batch size = {new_batch_size}", 'red')
                current_batch_size = new_batch_size

        cprint("❌ Batch size became too small. Stopping execution.", 'red')

    def get_accounts_data(self):
        """Create dicts with account info."""
        accounts_data = []
        for account in zip(ACCOUNT_NAMES, PRIVATE_KEYS, PROXIES):
            proxy = account[2] if GeneralSettings.USE_PROXY else False
            accounts_data.append({
                'account_name': account[0],
                'evm_private_key': account[1],
                'network': HyperTestnetRPC,
                'proxy': proxy
            })
        return accounts_data

    def chunk_list(self, lst, chunk_size):
        """Yield sublists of size chunk_size."""
        for i in range(0, len(lst), chunk_size):
            yield lst[i:i+chunk_size]

    async def process_batch(self, batch_accounts, batch_number):
        """Process a single batch and return results."""
        cprint(f"➡️  Processing batch #{batch_number}: {len(batch_accounts)} wallets.", 'cyan')
        clients = [Client(acc_data) for acc_data in batch_accounts]
        max_retries = 15
        retry_delay = 3

        try:
            calls, calls_info = self.create_calls_for_clients(clients)
            
            for attempt in range(max_retries):
                try:
                    # Randomly select a client for each attempt
                    random_client = clients[random.randint(0, len(clients) - 1)]
                    raw_results = await self.perform_multicall(random_client.w3, calls)

                    if raw_results is not None:
                        batch_wallets_data = self.parse_results(
                            clients, batch_accounts, calls_info, raw_results, (batch_number-1) * self.batch_size
                            )
                        return batch_wallets_data
                    
                except Exception as e:
                    cprint(f"⚠️ Attempt {attempt + 1} for #{batch_number} batch failed. Retrying...", 'yellow')
                
                if attempt < max_retries - 1:
                    await asyncio.sleep(retry_delay)
            
            raise ValueError(f"All {max_retries} attempts to process batch {batch_number} failed")

        finally:
            pass
            # close_tasks = [client.session.close() for client in clients]
            # await asyncio.gather(*close_tasks)

    def create_calls_for_clients(self, clients):
        """Build calls and info."""
        calls = []
        calls_info = []

        for wallet_index, client in enumerate(clients):
            address = client.address
            # getEthBalance
            calls.append('getEthBalance')
            calls.append(
                client.get_contract(MULTICALL3_CONTRACTS['HyperTestnet'], MULTICALL3_ABI).functions.getEthBalance(address)
            )
            calls_info.append({
                'wallet_index': wallet_index,
                'column_name': 'HYPE',
                'decimals': 18
            })

            # Token balances
            for token_name, token_address in TOKENS_PER_CHAIN['HyperTestnet'].items():
                if token_name != 'HYPE':
                    balance_call = client.get_contract(token_address).functions.balanceOf(address)
                    decimals_call = client.get_contract(token_address).functions.decimals()
                    calls.append(balance_call)
                    calls.append(decimals_call)
                    calls_info.append({
                        'wallet_index': wallet_index,
                        'column_name': token_name,
                        'is_token_balance': True
                    })
            
            # NFT balances
            # for nft_name, nft_address in NFTS_PER_CHAIN['HyperTestnet'].items():
            #     balance_call = client.get_contract(nft_address).functions.balanceOf(address)
            #     calls.append(balance_call)
            #     calls_info.append({
            #         'wallet_index': wallet_index,
            #         'column_name': nft_name,
            #         'decimals': 0,
            #     })

            # # Additional contract calls
            # contracts_calls = {
            #     'bend_deposit': (
            #         STATION_CONTRACTS['bend_rewards'],
            #         'balanceOf',
            #         [address],
            #         STATION_ABI['bend_rewards'],
            #         '$VDHONEY Bend',
            #         18
            #     ),
            #     'berps_deposit': (
            #         STATION_CONTRACTS['berps_rewards'],
            #         'balanceOf',
            #         [address],
            #         STATION_ABI['berps_rewards'],
            #         '$BHONEY Berps',
            #         18
            #     ),
            #     'bex_usdc_deposit': (
            #         STATION_CONTRACTS['bex_usdc_rewards'],
            #         'balanceOf',
            #         [address],
            #         STATION_ABI['bex_rewards'],
            #         '$HONEY-USDC Bex',
            #         18
            #     ),
            #     'bex_wbera_deposit': (
            #         STATION_CONTRACTS['bex_wbera_rewards'],
            #         'balanceOf',
            #         [address],
            #         STATION_ABI['bex_rewards'],
            #         '$HONEY-WBERA Bex',
            #         18
            #     ),
            #     'kodiak_yeet_deposit': (
            #         STATION_CONTRACTS['kodiak_yeet_rewards'],
            #         'balanceOf',
            #         [address],
            #         STATION_ABI['bex_rewards'],
            #         '$YEET-WBERA Kodiak',
            #         18
            #     ),
            #     'kodiak_ibgt_deposit': (
            #         STATION_CONTRACTS['kodiak_ibgt_rewards'],
            #         'balanceOf',
            #         [address],
            #         STATION_ABI['bex_rewards'],
            #         '$iBGT-WBERA Kodiak',
            #         18
            #     ),
            #     'ooga_booga': (
            #         OOGA_BOOGA_CONTRACTS['oogaticket'],
            #         'balanceOf',
            #         [address],
            #         OOGA_BOOGA_ABI['oogaticket'],
            #         'OOGA BOOGA',
            #         0
            #     ),
            #     'beranames': (
            #         BERANAMES_CONTRACTS['BaseRegistrar'],
            #         'balanceOf',
            #         [address],
            #         ERC20_ABI,
            #         'BERANAMES',
            #         0
            #     )
            # }

            # for call_name, call_data in contracts_calls.items():
            #     contract_address, func_name, args, abi, column_label, decimals = call_data
            #     call_obj = client.get_contract(contract_address, abi).functions[func_name](*args)
            #     calls.append(call_obj)
            #     calls_info.append({
            #         'wallet_index': wallet_index,
            #         'column_name': column_label,
            #         'decimals': decimals,
            #         'call_name': call_name
            #     })

        return calls, calls_info

    async def perform_multicall(self, w3, calls):
        """Run multicall and return raw results or None."""
        multicall = Multicall3(w3=w3, multicall_address=MULTICALL3_CONTRACTS['HyperTestnet'])
        try:
            raw_results = await multicall.aggregate3(*calls)
            return raw_results
        except Exception as e:
            raise e

    def parse_results(self, clients, accounts_data, calls_info, raw_results, batch_start_index):
        """Parse raw results into wallet data."""
        wallets_data = []
        for i in range(len(clients)):
            wallets_data.append({
                '#': batch_start_index + i + 1,  # Adjust the wallet number based on the batch start index
                'Account Name': accounts_data[i]['account_name']
            })

        current_result_idx = 0
        for info in calls_info:
            w_index = info['wallet_index']
            col_name = info['column_name']
            decimals = info.get('decimals', 18)

            if info.get('is_token_balance'):
                raw_balance = raw_results[current_result_idx]
                raw_decimals = raw_results[current_result_idx + 1]
                current_result_idx += 2

                if raw_balance is None:
                    raw_balance = 0
                if raw_decimals is None:
                    raw_decimals = 18

                val = clients[w_index].from_wei(raw_balance, raw_decimals)
                # Convert to float if val is a decimal
                val = float(val) if isinstance(val, decimal.Decimal) else val
                wallets_data[w_index][col_name] = val if val != 0 else 0
            else:
                raw_value = raw_results[current_result_idx]
                current_result_idx += 1

                if raw_value is None:
                    raw_value = 0

                value_formatted = clients[w_index].from_wei(raw_value, decimals)
                # Convert to float if value_formatted is a decimal
                value_formatted = (
                    float(value_formatted) 
                    if isinstance(value_formatted, decimal.Decimal) 
                    else value_formatted
                )
                wallets_data[w_index][col_name] = value_formatted if value_formatted != 0 else 0

        return wallets_data
    
    @staticmethod
    async def save_to_excel(wallets_data):
        """Save data to fast_wallets_stats.xlsx with totals."""
        # Round floats to 4 decimal places or format as needed
        def format_value(value):
            if isinstance(value, float):
                if value == 0:
                    return 0.0
                elif abs(value) < 1:
                    # Format to 4 significant digits after leading zeros
                    return float(f'{value:.4g}')
                else:
                    return round(value, 4)  # Round to 4 decimal places
            return value

        # Apply formatting to all data
        formatted_data = [{k: format_value(v) for k, v in row.items()} for row in wallets_data]

        xlsx_data = pd.DataFrame(formatted_data)
        directory = './data/accounts_stats/'
        if not os.path.exists(directory):
            os.makedirs(directory)
        excel_path = os.path.join(directory, 'fast_wallets_stats.xlsx')

        xlsx_data.to_excel(excel_path, index=False)
        workbook = load_workbook(excel_path)
        worksheet = workbook.active

        # Add totals row
        last_row = worksheet.max_row
        for col in range(3, worksheet.max_column + 1):
            col_letter = get_column_letter(col)
            sum_formula = f'=SUM({col_letter}2:{col_letter}{last_row})'
            sum_cell = worksheet.cell(row=last_row + 1, column=col, value=sum_formula)
            sum_cell.font = Font(bold=True)

        total_label_cell = worksheet.cell(row=last_row + 1, column=2, value="Total")
        total_label_cell.font = Font(bold=True)

        # Format columns
        for column in worksheet.columns:
            max_length = 0
            col_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
                cell.alignment = Alignment(horizontal='center', vertical='center')
            worksheet.column_dimensions[col_letter].width = max_length + 4

        workbook.save(excel_path)
        cprint('✅ Data successfully saved to /data/accounts_stats/fast_wallets_stats.xlsx', 'light_green', attrs=["blink"])
